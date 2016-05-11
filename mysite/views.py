#coding=utf-8
import datetime
from django.shortcuts import render_to_response
from django.http import HttpResponse,HttpResponseRedirect
from django.template import RequestContext
import sqlite3
import openpyxl
from openpyxl.cell import get_column_letter

import sys
# 默认编码改成utf8,不然会报错
reload(sys)
sys.setdefaultencoding("utf-8")

# 查询数据库取得第pn页的剧集信息
def getShow(pn):
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    csr.execute("select count(*) from sqlDemo")
    cnt = int(csr.fetchall()[0][0])
    offset = int(pn) * 10 - 10
    csr.execute("select * from sqlDemo order by time desc, ctime desc limit 10 offset ?", (offset,))
    items = csr.fetchall()
    return items, cnt

# 查询数据库取得第pn页的用户信息
def getShowUser(pn):
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    csr.execute("select count(*) from user")
    cnt = int(csr.fetchall()[0][0])
    offset = int(pn) * 10 - 10
    csr.execute("select * from user order by uid limit 10 offset ?", (offset,))
    items = csr.fetchall()
    return items, cnt

# 查询数据库取得UID为uid的用户第pn页的订阅信息
def getMyOrder(uid, pn):
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    csr.execute("select count(*) from od where uid=?", (uid,))
    cnt = int(csr.fetchall()[0][0])
    offset = int(pn) * 10 - 10
    csr.execute("select * from sqlDemo, od where sqlDemo.did=od.did and od.uid=? order by time desc, ctime desc limit 10 offset ?;", (uid, offset,))
    items = csr.fetchall()
    return items, cnt

# 查询数据库取得UID为uid的用户的所有订阅剧集ID
def getOrder(uid):
    odr = []
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    csr.execute("select * from od where uid=?", (uid,))
    items = csr.fetchall()
    for item in items:
        odr.append(item["did"])
    return odr

# 取得DID为did的剧集信息
def getOneItem(did):
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    csr.execute("select * from sqlDemo where did=?", (did,))
    items = csr.fetchall()
    ret = None
    for item in items:
        ret = item
    return ret

# 取得热门排行,口碑排行信息
def getTopHigh():
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    num = 5
    csr.execute("select * from sqlDemo order by rnum desc limit ?", (num,))
    top_items = csr.fetchall()
    csr.execute("select * from sqlDemo order by rate desc limit ?", (num,))
    high_items = csr.fetchall()
    return top_items, high_items

# 取得DID为did的剧集的评论相关信息
def getCmt(uid, did):
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    csr.execute("select user.uname, comment.* from user, comment where user.uid=comment.uid and did=?", (did,))
    cnt = 0
    cmts = csr.fetchall()
    # 获取评论数
    for cmt in cmts:
        cnt += 1
    csr.execute("select * from comment where did=?", (did,))
    dcmted = "no"
    dcmt = csr.fetchall()
    # 当前用户是否评论过该剧集
    if dcmt:
        dcmted = "yes"
    csr.execute("select * from comment where did=? and uid=?", (did,uid))
    mcmt = csr.fetchall()
    cmted = "no"
    if mcmt:
        cmted = "yes"

    return cnt, cmts, cmted, dcmted

# 获取替换用户信息
def getUserinfo(uid):
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    csr.execute("select * from user where uid=?", (uid,))
    uinfo = None
    items = csr.fetchall()
    for item in items:
        uinfo = item
    # 获取订阅数
    csr.execute("select count(*) from od where uid=?", (uid,))
    ocnt = 0
    items = csr.fetchall()
    for item in items:
        ocnt = item[0]

    return uinfo, ocnt

# 计算底部翻页组件需要展示的页码范围
def getRange(pn, pcnt):
    pn = int(pn)
    pcnt = int(pcnt)
    if pcnt <= 10:
        return 1, pcnt
    left = pn-4
    right = pn+5
    if pn - 4 < 1:
        left = 1
        right = 10
    elif pn + 5 > pcnt:
        left = pcnt - 9
        right = pcnt
    return left, right

# 欢迎页面(不用)
def welcome(req):
    return render_to_response('welcome.html')

# 返回登录页面
def login(req):
    # 已登录则直接跳到资源库
    if req.COOKIES.get("uid"):
        response = HttpResponseRedirect('show?pn=1')
        return response
    return render_to_response('login.html',{},context_instance=RequestContext(req))

# 返回注册页面
def register(req):
    return render_to_response('register.html',{},context_instance=RequestContext(req))

# 返回剧集信息页面
def cmt(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid=req.COOKIES.get("uid")
    did=req.GET["did"]
    edit = req.GET.get("edit")
    if not edit:
        edit = "0"
    auth = req.COOKIES.get("auth")
    top_items, high_items = getTopHigh()
    citem = getOneItem(did)
    ccnt, cmts, cmted, dcmted = getCmt(uid, did)
    # uid必须强制转换为int,不然删除评论逻辑会有问题
    dic = {"edit":edit, "auth":auth, "citem":citem, "username":username, "h_items":high_items, "t_items":top_items,
           "dcmted":dcmted, "uid":int(uid), "did":did, "ccnt":ccnt, "cmts":cmts, "cmted":cmted}

    return render_to_response('cmt.html',dic,context_instance=RequestContext(req))

# 返回资源库页面
def show(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    pn = req.GET.get("pn")
    if not pn:
        pn = 1
    auth = req.COOKIES.get("auth")
    pn = int(pn)
    items, cnt = getShow(pn)
    odr = getOrder(uid)
    top_items, high_items = getTopHigh()
    # 向上取整
    pcnt = (cnt + 9) / 10
    if pcnt == 0:
        pcnt = 1
    left, right = getRange(pn, pcnt)
    lim = []
    for i in range(left, right+1):
        lim.append(i)
    dic = {"auth":auth, "act":"show", "pn":int(pn), "uid":uid, "items":items, "username":username, "pcnt":int(pcnt), "odr":odr,
           "h_items":high_items, "t_items":top_items, "ppre":int(pn)-1, "pnxt":int(pn)+1, "lim":lim}
    return render_to_response('show.html',dic,context_instance=RequestContext(req))

# 返回用户信息页面
def showuser(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    pn = req.GET["pn"]
    auth = req.COOKIES.get("auth")
    pn = int(pn)
    items, cnt = getShowUser(pn)
    top_items, high_items = getTopHigh()
    # 向上取整
    pcnt = (cnt + 9) / 10
    if pcnt == 0:
        pcnt = 1
    left, right = getRange(pn, pcnt)
    lim = []
    for i in range(left, right+1):
        lim.append(i)
    dic = {"auth":auth, "act":"showuser", "pn":int(pn), "uid":uid, "items":items, "username":username, "pcnt":int(pcnt),
           "h_items":high_items, "t_items":top_items, "ppre":int(pn)-1, "pnxt":int(pn)+1, "lim":lim}
    return render_to_response('showuser.html',dic,context_instance=RequestContext(req))

# 返回剧集搜索页面
def search(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    auth = req.COOKIES.get("auth")

    top_items, high_items = getTopHigh()

    dic = {"act":"search",  "uid":uid, "username":username,"h_items":high_items, "t_items":top_items, "auth":auth}
    # 开始新一次搜索,要把之前搜索生成的cookies清掉
    response = render_to_response('search.html',dic,context_instance=RequestContext(req))
    response.delete_cookie("keyword")
    response.delete_cookie("tfrom")
    response.delete_cookie("tto")
    for i in range(4):
        tmp = req.COOKIES.get("src"+str(i))
        if tmp:
            response.delete_cookie("src"+str(i))

    return response

# 返回添加资源页面
def additem(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    auth = req.COOKIES.get("auth")
    add = req.GET.get("add")

    # 上一次添加失败,页面顶部应有警告框,故要把此参数传给模板
    if not add:
        add = "no"

    top_items, high_items = getTopHigh()

    dic = {"act":"addone",  "uid":uid, "username":username,"h_items":high_items,
           "t_items":top_items, "auth":auth, "add":add}
    response = render_to_response('additem.html',dic,context_instance=RequestContext(req))

    return response

# 返回用户搜索页面
def usearch(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    auth = req.COOKIES.get("auth")
    top_items, high_items = getTopHigh()
    dic = {"act":"usearch",  "uid":uid, "username":username,"h_items":high_items, "t_items":top_items,
           "auth":auth}
    # 开始新一次搜索,要吧之前搜索生成的cookies清掉
    response = render_to_response('usearch.html',dic,context_instance=RequestContext(req))
    response.delete_cookie("keyuname")
    response.delete_cookie("keyuid")
    return response

# 返回我的订阅页面
def myorder(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    pn = req.GET["pn"]
    auth = req.COOKIES.get("auth")
    pn = int(pn)
    items, cnt = getMyOrder(uid, pn)
    odr = getOrder(uid)
    top_items, high_items = getTopHigh()
    # 向上取整
    pcnt = (cnt + 9) / 10
    if pcnt == 0:
        pcnt = 1
    left, right = getRange(pn, pcnt)
    lim = []
    for i in range(left, right+1):
        lim.append(i)
    dic = {"auth":auth, "act":"myorder", "pn":int(pn), "uid":uid, "items":items, "username":username, "pcnt":int(pcnt), "odr":odr,
           "h_items":high_items, "t_items":top_items, "ppre":int(pn)-1, "pnxt":int(pn)+1, "lim":lim}
    return render_to_response('show.html',dic,context_instance=RequestContext(req))

# 返回用户信息页面
def uinfo(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.GET.get("uid")
    suid = req.GET.get("uid")
    err = req.GET.get("err")
    # 修改状态,传给模板
    modstat = req.GET.get("modstat")
    if modstat:
        modstat = str(modstat)
    else:
        modstat = ""
    # 错误状态,传给模板
    if err:
        err = str(err)
    else:
        err = ""
    auth = req.COOKIES.get("auth")
    top_items, high_items = getTopHigh()
    uinfo, ocnt = getUserinfo(uid)
    print uinfo
    dic = {"uid":uid, "username":username, "t_items":top_items, "h_items":high_items,
          "uinfo": uinfo, "ocnt":ocnt, "auth":auth, "suid":suid, "err":err, "modstat":modstat}
    return render_to_response('uinfo.html',dic,context_instance=RequestContext(req))

# 进行注册
def doreg(req):
    username = req.POST.get('username')
    email = req.POST.get('email')
    password = req.POST.get('password')
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    csr.execute("select * from user where uname=?", (username,))
    recs = csr.fetchall()
    if recs:
        return render_to_response('register.html', {"regstat":"fail", "str":"用户名已存在"},context_instance=RequestContext(req))
    csr.execute("select * from user where email=?", (email,))
    recs = csr.fetchall()
    if recs:
        return render_to_response('register.html', {"regstat":"fail", "str":"该邮箱已注册"},context_instance=RequestContext(req))
    conn.execute("insert into user(uname, pwd, email, auth) values(?, ?, ?, ?);", (username, password, email, 0,))
    conn.commit()
    return render_to_response('login.html',{"regstat":"ok"},context_instance=RequestContext(req))
    # str =  "<div class=\"alert alert-error alert-dismissable\">" \
    #        "<button type=\"button\" class=\"close\" data-dismiss=\"alert\" onclick" \
    #        "=\"cancel()\">&times;</button><strong>注册成功</strong></div>"
    # return render_to_response('login.html', {'str':str}, context_instance=RequestContext(req))

# 注册成功逻辑
def regok(req):
    return render_to_response('login.html',{"regstat":"ok"},context_instance=RequestContext(req))

# 登录逻辑
def dolog(req):
    if req.method == 'POST':
        username = req.POST.get('username')
        password = req.POST.get('password')
        conn = sqlite3.connect("/tmpdb/newdb.db")
        conn.row_factory = sqlite3.Row
        csr = conn.cursor()
        csr.execute("select uid, pwd, uname, auth from user where uname=?", (username,))
        recs = csr.fetchall()
        if recs:
            for rec in recs:
                if rec["pwd"] == password:
                    response = HttpResponseRedirect('show?pn=1')
                    response.set_cookie('username', rec["uname"], 3600)
                    response.set_cookie('uid', rec["uid"], 3600)
                    response.set_cookie('auth', rec["auth"], 3600)
                    return response
        return render_to_response("login.html", {"logstat":"fail", "str":"用户名或密码不正确"},context_instance=RequestContext(req))

    return render_to_response('show.html',{},context_instance=RequestContext(req))

# 退出逻辑
def logout(req):

    # 清除cookie并返回登录页
    response = render_to_response("login.html",{},context_instance=RequestContext(req))
    response.delete_cookie('username')
    response.delete_cookie('uid')

    return response

# 处理订阅ajax请求,更新数据库
def order(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    uid=req.GET["uid"]
    did=req.GET["did"]
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.execute("insert into od(uid, did) values(?, ?);", (uid, did,))
    conn.commit()
    print "done"

    return

# 处理退订ajax请求,更新数据库
def unorder(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    uid=req.GET["uid"]
    did=req.GET["did"]
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.execute("delete from od where uid=? and did=?;", (uid, did,))
    conn.commit()

    return

# 处理评论
def docmt(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    uid = req.GET["uid"]
    did = req.GET["did"]
    rate = float(req.POST.get("result"))
    txt = req.POST.get("cmtarea")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    mydate = datetime.datetime.now().strftime("%Y-%m-%d %X")

    conn.execute("insert into comment(uid, did, cmt, rate, date) values(?,?,?,?,?)",
                 (uid, did, txt, rate, mydate))
    conn.commit()
    csr.execute("select * from sqlDemo where did=?", (did,))
    recs = csr.fetchall()
    ditem = None
    for rec in recs:
        ditem = rec
    old_rnum = float(ditem["rnum"])
    old_rate = float(ditem["rate"])
    new_rnum = round(old_rnum + 1.0, 2)
    new_rate = round((old_rate * old_rnum + rate) / new_rnum, 2)
    conn.execute("update sqlDemo set rnum=?, rate=? where did=?", (new_rnum, new_rate, did))
    conn.commit()
    response = HttpResponseRedirect('cmt?uid='+str(uid)+'&did='+str(did))
    return response

# 执行剧集搜索返回结果页面
def dosearch(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    kwd = req.POST.get("keyword")
    tfr = req.POST.get("tfrom")
    tto = req.POST.get("tto")
    srcs = req.REQUEST.getlist("source")
    auth = req.COOKIES.get("auth")
    if not kwd:
        # 可能是翻页,故应从 cookie 中取得相关参数
        kwd = req.COOKIES.get("keyword")
        tfr = req.COOKIES.get("tfrom")
        tto = req.COOKIES.get("tto")
        srcs = []
        for i in range(4):
            src = req.COOKIES.get("src"+str(i))
            if src is not None:
                srcs.append(str(src))
    pn = req.GET["pn"]
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    offset = int(pn) * 10 - 10
    recs = None
    cnt = 0
    # 没有关键词则默认搜索所有剧集信息
    if not kwd:
        kwd = ""
    # 生成搜索sql字符串
    sqlstr = "select * from sqlDemo where title like '%"+kwd+"%'"
    if tfr:
        sqlstr += " and time > '" + tfr + "' "
    if tto:
        sqlstr += " and time < '" + tto + "' "
    if srcs:
        sqlstr += " and ( "
        first = True
        for src in srcs:
            if first:
                first = False
            else:
                sqlstr += " or "
            sqlstr += " src = '" + src + "' "
        # sqlstr用于计算cnt,sqlstrlim用于返回结果
        sqlstr += " ) order by time desc, ctime desc "
        print sqlstr
        sqlstrlim = sqlstr + "limit 10 offset " + str(offset)
        csr.execute(sqlstr)
        # all records
        recall = csr.fetchall()
        if recall:
            for eachrec in recall:
                cnt += 1
        csr.execute(sqlstrlim)
        recs = csr.fetchall()
    # 总页数 向上取整
    pcnt = (cnt + 9) / 10
    if pcnt == 0:
        pcnt = 1
    top_items, high_items = getTopHigh()
    odr = getOrder(uid)
    left, right = getRange(pn, pcnt)
    lim = []
    for i in range(left, right+1):
        lim.append(i)
    dic = {"act":"dosearch", "pn":int(pn), "uid":uid, "items":recs, "username":username, "pcnt":int(pcnt), "odr":odr,
           "h_items":high_items, "t_items":top_items, "ppre":int(pn)-1, "pnxt":int(pn)+1, "lim":lim, "auth":auth}
    response = render_to_response('show.html',dic,context_instance=RequestContext(req))
    response.set_cookie("keyword", kwd)
    response.set_cookie("tfrom", tfr)
    response.set_cookie("tto", tto)
    i = 0
    for src in srcs:
        response.set_cookie("src"+str(i), src)
        i += 1
    return response

# 执行用户搜索返回结果页面
def dousearch(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    auth = req.COOKIES.get("auth")
    keyuid = req.POST.get("keyuid")
    # 搜索类型: 按用户名 or 按UID
    stype = req.POST.get("stype")
    if not keyuid:
        keyuid = req.COOKIES.get("keyuid")
        stype = req.COOKIES.get("stype")
    pn = req.GET.get("pn")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    csr = conn.cursor()
    conn.row_factory = sqlite3.Row
    offset = int(pn) * 10 - 10
    # recs = None
    sqlstr = "select * from user "
    if stype == "byuid":
        sqlstr += " where uid = " + str(keyuid)
    else:
        sqlstr += " where uname like '%"+str(keyuid)+"%'"
    # sqlstr用于计算cnt,sqlstrlim用于返回结果
    sqlstrlim = sqlstr + " limit 10 offset " + str(offset)
    csr.execute(sqlstr)
    items = csr.fetchall()
    cnt = 0
    for item in items:
        cnt += 1
    pcnt = (cnt + 9) / 10
    if pcnt == 0:
        pcnt = 1
    csr.execute(sqlstrlim)
    items = csr.fetchall()
    top_items, high_items = getTopHigh()
    left, right = getRange(pn, pcnt)
    lim = []
    for i in range(left, right+1):
        lim.append(i)
    dic = {"auth":auth, "act":"dousearch", "pn":int(pn), "uid":uid, "items":items, "username":username, "pcnt":int(pcnt),
           "h_items":high_items, "t_items":top_items, "ppre":int(pn)-1, "pnxt":int(pn)+1, "lim":lim}
    response = render_to_response('showuser.html',dic,context_instance=RequestContext(req))
    if keyuid:
        response.set_cookie("keyuid", keyuid)
    if stype:
        response.set_cookie("stype", stype)

    return response

# 修改邮箱(不用,改用moduinfo)
def modemail(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    uid=req.GET["uid"]
    newemail = req.POST.get("newemail")
    newuname = req.POST.get("newuname")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    csr.execute("select * from user where uname=?", (newuname,))
    unames = csr.fetchall()
    if unames:
        response = HttpResponseRedirect("uinfo?uid="+uid+"&err=exuname")
        return response
    csr.execute("select * from user where email=?", (newemail,))
    emails = csr.fetchall()
    if emails:
        response = HttpResponseRedirect("uinfo?uid="+uid+"&err=exemail")
        return response
    sqlstr = None
    if newemail:
        sqlstr = "update user set email='"+newemail+"'"
        if newuname:
            sqlstr += ",uname='"+newuname+"'"
    elif newuname:
        sqlstr = "update user set uname='"+newuname+"'"
    if sqlstr != None:
        sqlstr += " where uid="+uid
        conn.execute(sqlstr)
        conn.commit()
    response = HttpResponseRedirect("uinfo?uid="+uid)
    return response

# 修改剧集信息
def modinfo(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    uid = req.COOKIES.get("uid")
    did = req.GET.get("did")
    newtit = req.POST.get("newtit")
    newlink = req.POST.get("newlink")
    sqlstr = None
    if newtit:
        sqlstr = "update sqlDemo set title='"+newtit+"'"
        if  newlink:
            sqlstr += ", link='"+newlink+"'"
    elif newlink:
        sqlstr = "update sqlDemo set link='"+newlink+"'"
    if sqlstr:
        sqlstr += " where did="+did
        conn = sqlite3.connect("/tmpdb/newdb.db")
        conn.execute(sqlstr)
        conn.commit()
    response = HttpResponseRedirect("cmt?uid="+str(uid)+"&did="+str(did))
    return response

# 返回修改用户信息页面
def moduinfo(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    username=req.COOKIES.get("username")
    uid = req.COOKIES.get("uid")
    auth = req.COOKIES.get("auth")
    muid = req.GET.get("muid")
    err = req.GET.get("err")
    uid = str(uid)
    muid = str(muid)
    showpwd = "no"
    if uid == muid:
        showpwd = "yes"
    if not err:
        err = ""

    top_items, high_items = getTopHigh()

    dic = {"act":"search",  "uid":str(uid), "username":username,"h_items":high_items, "t_items":top_items,
           "auth":auth, "muid":str(muid), "showpwd":showpwd, "err":err}

    response = render_to_response('moduinfo.html',dic,context_instance=RequestContext(req))
    return response

# 访问数据库修改用户信息
def domodu(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    muid=req.GET.get("muid")
    oldpwd=req.POST.get("oldpwd")
    newpwd=req.POST.get("newpwd")
    newemail=req.POST.get("newemail")
    muid = str(muid)
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    if oldpwd:
        csr.execute("select pwd from user where uid=?", (muid,))
        dbpwd = csr.fetchall()[0][0]
        print oldpwd, dbpwd
        if oldpwd != dbpwd:
            response = HttpResponseRedirect("moduinfo?muid="+str(muid)+"&err=pwdwrng")
            return response
        if newemail:
            csr.execute("select * from user where email=?", (newemail,))
            emails = csr.fetchall()
            if emails:
                response = HttpResponseRedirect("moduinfo?muid="+muid+"&err=exemail")
                return response
            conn.execute("update user set email=? where uid=?", (newemail, muid))
        if newpwd:
            conn.execute("update user set pwd=? where uid=?", (newpwd, muid))
    conn.commit()
    response = HttpResponseRedirect("uinfo?uid="+muid+"&modstat=ok")
    return response

# 访问数据库添加资源
def doadditem(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    atitle = req.POST.get("atitle")
    alink = req.POST.get("alink")
    atime = req.POST.get("atime")
    asrc = req.POST.get("asrc")
    atitle = str(atitle)
    alink = str(alink)
    if not atime:
        atime = datetime.datetime.now().strftime("%Y-%m-%d %X")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.execute("insert into sqlDemo(title, link, src, img, time, ctime, rate, rnum) values(?,?,?,?,?,?,?,?)",
                 (atitle, alink, asrc, "/static/han.png", atime, atime, 0, 0) )
    conn.commit()
    response = HttpResponseRedirect("additem?add=ok")
    return response

# 删除剧集信息
def delinfo(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    did = req.GET.get("did")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.execute("delete from sqlDemo where did=?", (did,))
    conn.commit()
    response = HttpResponseRedirect("show?pn=1")
    return response

# 删除用户信息
def deluser(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    duid = req.GET.get("duid")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    # 删除用户之前要把他的评论和订阅删掉
    csr.execute("select * from comment where uid=?", (duid,))
    dcmts = csr.fetchall()
    for dcmt in dcmts:
        rate = dcmt["rate"]
        did = dcmt["did"]
        csr.execute("select * from sqlDemo where did=?", (did,))
        recs = csr.fetchall()
        ditem = None
        for rec in recs:
            ditem = rec
        old_rnum = float(ditem["rnum"])
        old_rate = float(ditem["rate"])
        new_rnum = round(old_rnum - 1.0, 2)
        if new_rnum == 0.0:
            new_rate = 0
        else:
            new_rate = round((old_rate * old_rnum - rate) / new_rnum, 2)
        conn.execute("update sqlDemo set rnum=?, rate=? where did=?", (new_rnum, new_rate, did))
        conn.execute("delete from comment where uid=? and did=?", (duid, did))
    conn.execute("delete from od where uid=?", (duid,))
    conn.execute("delete from user where uid=?", (duid,))
    conn.commit()
    response = HttpResponseRedirect("showuser?pn=1")
    return response

# 删除评论
def delcmt(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    did = req.GET.get("did")
    uid = req.GET.get("uid")
    myuid = req.COOKIES.get("uid")
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    csr.execute("select * from comment where uid=? and did=?", (uid, did,))
    ocmts = csr.fetchall()
    pcmt = None
    for ocmt in ocmts:
        pcmt = ocmt
    rate = pcmt["rate"]
    if not rate:
        rate = float(5.0)
    conn.execute("delete from comment where uid=? and did=?", (uid, did,))
    conn.commit()
    # 删除评论后要更新相应剧集信息的评分
    csr.execute("select * from sqlDemo where did=?", (did,))
    recs = csr.fetchall()
    ditem = None
    for rec in recs:
        ditem = rec
    old_rnum = float(ditem["rnum"])
    old_rate = float(ditem["rate"])
    new_rnum = round(old_rnum - 1.0, 2)
    if new_rnum == 0.0:
        new_rate = 0
    else:
        new_rate = round((old_rate * old_rnum - rate) / new_rnum, 2)
    conn.execute("update sqlDemo set rnum=?, rate=? where did=?", (new_rnum, new_rate, did))
    conn.commit()
    response = HttpResponseRedirect("cmt?uid="+str(myuid)+"&did="+str(did))
    return response

# 生成用户信息 excel
def userexcel(req):
    if not req.COOKIES.get("uid") or req.COOKIES.get("auth") != "1":
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=userinfo.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = u"用户列表"+datetime.datetime.now().strftime("%Y-%m-%d")

    row_num = 0

    columns = [
        (u"UID", 5),
        (u"用户名", 15),
        (u"密码", 15),
        (u"邮箱", 20),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    csr.execute("select * from user order by uid")
    items = csr.fetchall()
    for item in items:
        row_num += 1
        row = [
            item["uid"],
            item["uname"],
            item["pwd"],
            item["email"]
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

# 生成剧集信息 excel
def dramaexcel(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dramainfo.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = u"剧集信息"+datetime.datetime.now().strftime("%Y-%m-%d")

    row_num = 0

    columns = [
        (u"标题", 70),
        (u"更新时间", 21),
        (u"来源", 11),
        (u"评分", 5),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
    conn = sqlite3.connect("/tmpdb/newdb.db")
    conn.row_factory = sqlite3.Row
    csr = conn.cursor()
    csr.execute("select * from sqlDemo order by time desc, ctime desc")
    items = csr.fetchall()
    for item in items:
        row_num += 1
        row = [
            item["title"],
            item["time"],
            item["src"],
            item["rate"]
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

# 测试用
def test(req):
    if not req.COOKIES.get("uid"):
        return render_to_response('login.html',{},context_instance=RequestContext(req))
    return render_to_response('/static/test.txt')




